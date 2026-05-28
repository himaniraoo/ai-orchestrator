import uuid
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

ARTIFACTS_DIR = Path(__file__).parent.parent / "artifacts"
ARTIFACTS_DIR.mkdir(exist_ok=True)

# -------------------------------------------------------------------
# DESIGN TOKENS
# -------------------------------------------------------------------
COLOR_HEADER_BG  = "0D1B2A"
COLOR_HEADER_FG  = "FFFFFF"
COLOR_ACCENT     = "00A8E8"
COLOR_ROW_ALT    = "E8F4FD"
COLOR_ROW_NORMAL = "FFFFFF"
COLOR_BORDER     = "CCCCCC"


# -------------------------------------------------------------------
# STYLE HELPERS
# -------------------------------------------------------------------

def _header_font():
    return Font(bold=True, color=COLOR_HEADER_FG, size=11)

def _normal_font():
    return Font(size=10)

def _header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)

def _alt_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_ALT)

def _normal_fill():
    return PatternFill("solid", fgColor=COLOR_ROW_NORMAL)

def _thin_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row_num: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.border = _thin_border()
        cell.alignment = _center()


def _style_data_row(ws, row_num: int, num_cols: int, alternate: bool):
    fill = _alt_fill() if alternate else _normal_fill()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _normal_font()
        cell.fill = fill
        cell.border = _thin_border()
        cell.alignment = _left()


def _autofit_columns(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


# -------------------------------------------------------------------
# CLAIM VOLUME HELPER
# -------------------------------------------------------------------

def _get_relevant_claims(physician: dict, icd10_codes: list[str] = None) -> int:
    """
    Returns claim volume for selected ICD-10 codes only.
    Falls back to totalNSCLCClaims if no codes specified.
    """
    if not icd10_codes:
        return physician.get("totalNSCLCClaims", 0)

    selected = [c.upper() for c in icd10_codes]
    return sum(
        claims for code, claims in physician.get("icd10ClaimVolume", {}).items()
        if code.upper() in selected
    )


# -------------------------------------------------------------------
# SHEET BUILDERS
# -------------------------------------------------------------------

def _build_raw_data_sheet(wb: Workbook, physicians: list[dict], icd10_codes: list[str] = None):
    """
    Sheet 1 — Raw physician data table.
    Claims column reflects selected ICD-10 codes only if provided.
    Already filtered per preferences by the orchestrator.
    """
    ws = wb.active
    ws.title = "Raw Physician Data"

    # Dynamic header — reflects which codes are being measured
    claims_header = (
        f"{', '.join(icd10_codes)} Claims" if icd10_codes
        else "Total NSCLC Claims"
    )

    headers = [
        "NPI", "First Name", "Last Name", "Specialty",
        "Affiliation", "City", "State",
        claims_header, "Volume Tier", "Board Certified"
    ]

    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20

    for ri, p in enumerate(physicians, start=2):
        values = [
            p.get("npi", ""),
            p.get("firstName", ""),
            p.get("lastName", ""),
            p.get("specialty", ""),
            p.get("affiliation", ""),
            p.get("city", ""),
            p.get("state", ""),
            _get_relevant_claims(p, icd10_codes),
            p.get("volumeTier", "").replace("_", " ").title(),
            "Yes" if p.get("boardCertified") else "No",
        ]
        for ci, val in enumerate(values, start=1):
            ws.cell(row=ri, column=ci, value=val)
        _style_data_row(ws, ri, len(headers), alternate=(ri % 2 == 0))

    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _build_pivot_sheet(wb: Workbook, physicians: list[dict], icd10_codes: list[str] = None):
    """
    Sheet 2 — Pivot-style summary: claim volume by state x specialty.
    Rows = states, Columns = specialties, Values = selected ICD-10 claim volume.
    """
    ws = wb.create_sheet(title="State x Specialty Pivot")

    # Row 1 — metric label so reviewer knows what's being measured
    metric_label = (
        f"Claim Volume ({', '.join(icd10_codes)})" if icd10_codes
        else "Total NSCLC Claim Volume"
    )
    ws.cell(row=1, column=1, value=metric_label)
    ws.cell(row=1, column=1).font = Font(bold=True, italic=True, size=10, color="6B7280")

    states = sorted(set(p["state"] for p in physicians))
    specialties = sorted(set(p["specialty"] for p in physicians))

    pivot: dict[tuple, int] = defaultdict(int)
    state_totals: dict[str, int] = defaultdict(int)
    specialty_totals: dict[str, int] = defaultdict(int)

    for p in physicians:
        key = (p["state"], p["specialty"])
        relevant_claims = _get_relevant_claims(p, icd10_codes)
        pivot[key] += relevant_claims
        state_totals[p["state"]] += relevant_claims
        specialty_totals[p["specialty"]] += relevant_claims

    num_cols = len(specialties) + 2

    # Header row at row 2
    header_row = 2
    ws.cell(row=header_row, column=1, value="State")
    for ci, spec in enumerate(specialties, start=2):
        ws.cell(row=header_row, column=ci, value=spec)
    ws.cell(row=header_row, column=len(specialties) + 2, value="State Total")
    _style_header_row(ws, header_row, num_cols)
    ws.row_dimensions[header_row].height = 20

    # Data rows start at row 3
    for ri, state in enumerate(states, start=3):
        ws.cell(row=ri, column=1, value=state)
        for ci, spec in enumerate(specialties, start=2):
            val = pivot.get((state, spec), 0)
            ws.cell(row=ri, column=ci, value=val if val > 0 else "—")
        ws.cell(row=ri, column=len(specialties) + 2, value=state_totals[state])
        _style_data_row(ws, ri, num_cols, alternate=(ri % 2 == 0))

    # Totals row at bottom
    total_row = len(states) + 3
    ws.cell(row=total_row, column=1, value="Specialty Total")
    for ci, spec in enumerate(specialties, start=2):
        ws.cell(row=total_row, column=ci, value=specialty_totals[spec])
    ws.cell(row=total_row, column=len(specialties) + 2, value=sum(state_totals.values()))
    _style_header_row(ws, total_row, num_cols)

    ws.freeze_panes = "B3"
    _autofit_columns(ws)


def _build_icd10_sheet(wb: Workbook, physicians: list[dict], icd10_codes: list[str] = None):
    """
    Sheet 3 — ICD-10 code breakdown — physician count per SELECTED code only.
    If icd10_codes provided, only those codes are shown.
    If not provided, all codes in the dataset are shown.
    """
    ws = wb.create_sheet(title="ICD-10 Breakdown")

    selected_codes = [c.upper() for c in icd10_codes] if icd10_codes else None

    code_physician_count: dict[str, int] = defaultdict(int)
    code_total_claims: dict[str, int] = defaultdict(int)

    for p in physicians:
        for code, claims in p.get("icd10ClaimVolume", {}).items():
            if selected_codes and code.upper() not in selected_codes:
                continue
            code_physician_count[code] += 1
            code_total_claims[code] += claims

    headers = [
        "ICD-10 Code",
        "Physicians with Claims",
        "Total Claims Across Physicians",
        "Avg Claims per Physician"
    ]

    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20

    # Show selected codes in order — even if zero physicians
    # so reviewer sees the code was checked
    if selected_codes:
        sorted_codes = selected_codes
    else:
        sorted_codes = sorted(
            code_total_claims.keys(),
            key=lambda c: code_total_claims[c],
            reverse=True
        )

    for ri, code in enumerate(sorted_codes, start=2):
        count = code_physician_count.get(code, 0)
        total = code_total_claims.get(code, 0)
        avg = round(total / count, 1) if count > 0 else 0
        values = [code, count, total, avg]
        for ci, val in enumerate(values, start=1):
            ws.cell(row=ri, column=ci, value=val)
        _style_data_row(ws, ri, len(headers), alternate=(ri % 2 == 0))

    ws.freeze_panes = "A2"
    _autofit_columns(ws)


# -------------------------------------------------------------------
# MAIN ENTRY POINT
# -------------------------------------------------------------------

def run_excel_agent(
    analysis_type: str,
    physician_list: list[dict],
    dimensions: list[str] = None,
    icd10_codes: list[str] = None,
) -> dict:
    """
    Generates a real .xlsx workbook with 3 sheets and returns artifact metadata.
    Pure Python — no LLM needed for structured tabular output.
    Sheet 1: Raw physician data (filtered per preferences, code-specific claims)
    Sheet 2: Pivot summary — code-specific claim volume by state x specialty
    Sheet 3: ICD-10 breakdown — physician count per selected code only
    """

    if not physician_list:
        return {"error": "No physician data provided to Excel agent"}

    wb = Workbook()

    _build_raw_data_sheet(wb, physician_list, icd10_codes=icd10_codes)
    _build_pivot_sheet(wb, physician_list, icd10_codes=icd10_codes)
    _build_icd10_sheet(wb, physician_list, icd10_codes=icd10_codes)

    artifact_id = f"docnexus_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = ARTIFACTS_DIR / artifact_id
    wb.save(str(output_path))

    return {
        "status": "success",
        "artifact_id": artifact_id,
        "download_url": f"/artifacts/{artifact_id}",
        "sheet_count": 3,
        "physician_count": len(physician_list),
        "sheets": [
            "Raw Physician Data",
            "State x Specialty Pivot",
            "ICD-10 Breakdown"
        ]
    }